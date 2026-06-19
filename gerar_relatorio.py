"""
Gerador de relatório histórico de LPCOs.

Uso no servidor Hetzner:
  python gerar_relatorio.py                          # relatório completo
  python gerar_relatorio.py --de 2026-06-01          # a partir de junho
  python gerar_relatorio.py --de 2026-06-01 --ate 2026-06-19

O script faz duas coisas:
  1. Importa os eventos do arquivo lpco_monitor.log para a tabela `eventos`
     (migração única — eventos duplicados são ignorados).
  2. Gera relatorio_lpco_<de>_<ate>.xlsx com os dados formatados.

Dependência: openpyxl  (pip install openpyxl)
"""

import argparse
import json
import logging
import re
import sqlite3
import sys
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(levelname)s — %(message)s")
logger = logging.getLogger(__name__)

LOG_FILE = Path("lpco_monitor.log")
DB_PATH  = Path("lpco_monitor.db")

# Regex para linhas de payload completo
_LOG_RE = re.compile(
    r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ \[INFO\] webhook_receiver"
    r" — === PAYLOAD COMPLETO === "
    r"event=(?P<event>\S+) "
    r"destinatario-id=(?P<dest_id>\S+) "
    r"numeroLPCO=(?P<numero>\S+) "
    r"codigoModelo=(?P<modelo>\S+) "
    r"payload=(?P<payload>.+)$"
)

CERT_NE_OWNER_ID: str = ""  # carregado do .env abaixo


def _carregar_ne_owner() -> str:
    try:
        from config import config  # noqa: PLC0415
        return config.CERT_NE_OWNER_ID
    except Exception:
        return ""


def _importar_log(conn: sqlite3.Connection) -> int:
    """Lê lpco_monitor.log e insere eventos novos na tabela eventos. Retorna qtd importada."""
    if not LOG_FILE.exists():
        logger.warning("Arquivo de log '%s' não encontrado.", LOG_FILE)
        return 0

    importados = 0
    with open(LOG_FILE, encoding="utf-8", errors="replace") as f:
        for linha in f:
            m = _LOG_RE.match(linha.rstrip())
            if not m or m.group("event") != "talp-altsit-lpco-anu":
                continue

            ts        = m.group("ts")          # "2026-06-05 14:32:11"
            dest_id   = m.group("dest_id")
            numero    = m.group("numero")
            modelo    = m.group("modelo")
            json_str  = m.group("payload")

            try:
                payload = json.loads(json_str)
            except json.JSONDecodeError:
                continue

            nova_sit  = payload.get("novaSituacao", {})
            sit_id    = nova_sit.get("id", "").upper()
            sit_desc  = nova_sit.get("descricao", sit_id)
            justif    = payload.get("justificativa", "")
            cnpj_list = payload.get("cpfCnpj", [])
            data_ev   = payload.get("dataEvento", ts)
            regiao    = "NE" if (CERT_NE_OWNER_ID and dest_id == CERT_NE_OWNER_ID) else "SE"
            tipo      = f"Pesca {regiao}"

            # Evita duplicatas (mesma data_evento + numero + situacao)
            existe = conn.execute(
                "SELECT 1 FROM eventos WHERE data_evento=? AND numero_lpco=? AND situacao_id=?",
                (data_ev, numero, sit_id),
            ).fetchone()
            if existe:
                continue

            conn.execute(
                """
                INSERT INTO eventos
                    (data_evento, data_recebido, numero_lpco, codigo_modelo, tipo, regiao,
                     destinatario_id, situacao_id, situacao_desc, justificativa, cnpj_cpf, payload_json)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    data_ev, ts, numero, modelo, tipo, regiao,
                    dest_id, sit_id, sit_desc, justif,
                    ", ".join(cnpj_list),
                    json_str,
                ),
            )
            importados += 1

    conn.commit()
    return importados


def _gerar_excel(eventos: list[dict], caminho: Path) -> None:
    try:
        import openpyxl                             # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill, Alignment  # noqa: PLC0415
    except ImportError:
        logger.error("openpyxl não instalado. Execute: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eventos LPCO"

    cabecalhos = [
        "Data do Evento", "Data Recebido", "Número LPCO", "Código Modelo",
        "Tipo", "Região", "Destinatário ID", "Situação", "Descrição Situação",
        "Justificativa", "CNPJ/CPF Importador",
    ]
    chaves = [
        "data_evento", "data_recebido", "numero_lpco", "codigo_modelo",
        "tipo", "regiao", "destinatario_id", "situacao_id", "situacao_desc",
        "justificativa", "cnpj_cpf",
    ]

    # Cabeçalho
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, titulo in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Cores por situação
    _COR_SITUACAO = {
        "DEFERIDO":       "C6EFCE",   # verde claro
        "INDEFERIDO":     "FFC7CE",   # vermelho claro
        "EM_ANALISE":     "FFEB9C",   # amarelo
        "EM_VERIFICACAO": "DDEBF7",   # azul claro
    }

    for row_idx, ev in enumerate(eventos, start=2):
        sit = (ev.get("situacao_id") or "").upper()
        fill_color = _COR_SITUACAO.get(sit, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_color) if fill_color != "FFFFFF" else None

        for col_idx, chave in enumerate(chaves, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=ev.get(chave, ""))
            if row_fill:
                cell.fill = row_fill

    # Larguras de coluna
    larguras = [20, 20, 18, 14, 12, 8, 18, 16, 22, 40, 22]
    for col_idx, larg in enumerate(larguras, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = larg

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(caminho)
    logger.info("Relatório salvo em: %s  (%d linhas)", caminho, len(eventos))


def main() -> None:
    global CERT_NE_OWNER_ID
    CERT_NE_OWNER_ID = _carregar_ne_owner()

    parser = argparse.ArgumentParser(description="Gera relatório Excel de eventos LPCO.")
    parser.add_argument("--de",  default="", help="Data inicial YYYY-MM-DD")
    parser.add_argument("--ate", default="", help="Data final   YYYY-MM-DD")
    parser.add_argument("--sem-importar-log", action="store_true",
                        help="Pula a importação do log (usa só o que já está no banco)")
    args = parser.parse_args()

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row

        # Garante que a tabela existe
        conn.execute("""
            CREATE TABLE IF NOT EXISTS eventos (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                data_evento     TEXT,
                data_recebido   TEXT DEFAULT (datetime('now')),
                numero_lpco     TEXT,
                codigo_modelo   TEXT,
                tipo            TEXT,
                regiao          TEXT,
                destinatario_id TEXT,
                situacao_id     TEXT,
                situacao_desc   TEXT,
                justificativa   TEXT,
                cnpj_cpf        TEXT,
                payload_json    TEXT
            )
        """)
        conn.commit()

        if not args.sem_importar_log:
            n = _importar_log(conn)
            logger.info("Log importado: %d evento(s) novo(s) inserido(s) no banco.", n)

        # Lê eventos filtrados
        query = "SELECT * FROM eventos WHERE 1=1"
        params: list[str] = []
        if args.de:
            query += " AND data_evento >= ?"
            params.append(args.de)
        if args.ate:
            query += " AND data_evento <= ?"
            params.append(args.ate + " 23:59:59")
        query += " ORDER BY data_evento ASC"

        rows = conn.execute(query, params).fetchall()
        eventos = [dict(r) for r in rows]

    logger.info("Total de eventos encontrados: %d", len(eventos))
    if not eventos:
        logger.warning("Nenhum evento no período. Relatório não gerado.")
        return

    sufixo = f"{args.de or 'completo'}_{args.ate or 'hoje'}"
    caminho = Path(f"relatorio_lpco_{sufixo}.xlsx")
    _gerar_excel(eventos, caminho)


if __name__ == "__main__":
    main()
