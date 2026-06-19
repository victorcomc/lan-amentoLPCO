"""
Relatório semanal — módulo independente, não altera o sistema de webhooks.

Abas geradas no Excel:
  1. "Resumo por Cliente (LPCO)"   — totais por CNPJ: LPCOs, peso, produtos, destinos, situações
  2. "Detalhe LPCO"                — uma linha por LPCO, todos os campos disponíveis
  3. "Resumo Mercado (DUEs)"       — totais por exportador: DUEs, peso, FOB, destinos
  4. "Detalhe DUEs"                — uma linha por evento DUE recebido
  5. "Análise de Mercado"          — rankings: top exportadores, destinos, produtos, modais
  6. "Legenda"                     — referência de cores por situação

Agendamento padrão: toda sexta-feira às 14:00 (configurável em main.py).
"""

import logging
import time
from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any

from config import config
from database import listar_dues
from siscomex_client import SiscomexClient

logger = logging.getLogger(__name__)

_DELAY_ENTRE_CHAMADAS = 0.4
_TAMANHO_PAGINA = 50

_CORES_SITUACAO = {
    "DEFERIDO":         "C6EFCE",
    "INDEFERIDO":       "FFC7CE",
    "EM_ANALISE":       "FFEB9C",
    "EM_VERIFICACAO":   "DDEBF7",
    "CANCELADO":        "D9D9D9",
    "VENCIDO":          "F2DCDB",
    "SUSPENSAO":        "E2EFDA",
}

_MODELOS = {
    "E00061": "Pesca",
    "E00144": "Fruta",
}


# ---------------------------------------------------------------------------
# Descoberta e detalhe via API
# ---------------------------------------------------------------------------

def _paginar_lpcos(client: SiscomexClient, label: str) -> list[str]:
    numeros: list[str] = []
    pagina = 1
    while True:
        resultado = client.buscar_lpcos(pagina=pagina, tamanho=_TAMANHO_PAGINA)
        if not resultado.sucesso or not resultado.registros:
            if pagina == 1 and not resultado.sucesso:
                logger.warning("Falha na paginação %s: %s", label, resultado.erro)
            break
        for r in resultado.registros:
            if r.numero:
                numeros.append(r.numero)
        logger.info("%s: página %d — %d LPCOs (acumulado: %d)", label, pagina, len(resultado.registros), len(numeros))
        if len(resultado.registros) < _TAMANHO_PAGINA:
            break
        pagina += 1
        time.sleep(_DELAY_ENTRE_CHAMADAS)
    return numeros


def _detalhar_em_sessao(
    numeros: list[str],
    pfx_path: str, pfx_base64: str, pfx_password: str,
    label: str,
) -> dict[str, dict]:
    resultado: dict[str, dict] = {}
    if not (pfx_path or pfx_base64):
        return resultado
    try:
        with SiscomexClient(
            cert_pfx_path=pfx_path,
            cert_pfx_base64=pfx_base64,
            cert_pfx_password=pfx_password,
        ) as client:
            if not client.autenticar(config.WEBHOOK_ROLE_TYPE):
                logger.error("Autenticação %s falhou para relatório.", label)
                return resultado
            encontrados = _paginar_lpcos(client, label)
            pendentes = [n for n in encontrados if n not in numeros]
            logger.info("%s: %d encontrados, %d novos para detalhar.", label, len(encontrados), len(pendentes))
            for i, numero in enumerate(pendentes, start=1):
                try:
                    raw = client.detalhar_lpco(numero)
                    resultado[numero] = raw if isinstance(raw, dict) else {}
                except Exception as exc:
                    logger.debug("%s: detalhe %s — %s", label, numero, exc)
                    resultado[numero] = {}
                if i % 50 == 0:
                    logger.info("  %s: %d/%d detalhes obtidos.", label, i, len(pendentes))
                time.sleep(_DELAY_ENTRE_CHAMADAS)
    except Exception as exc:
        logger.error("Erro na sessão %s: %s", label, exc)
    return resultado


def _buscar_todos_detalhes() -> dict[str, dict]:
    detalhes = _detalhar_em_sessao(
        numeros=[],
        pfx_path=config.CERT_PFX_PATH, pfx_base64=config.CERT_PFX_BASE64,
        pfx_password=config.CERT_PFX_PASSWORD, label="SE",
    )
    if config.CERT_NE_PFX_BASE64 or config.CERT_NE_PFX_PATH:
        ne = _detalhar_em_sessao(
            numeros=list(detalhes.keys()),
            pfx_path=config.CERT_NE_PFX_PATH, pfx_base64=config.CERT_NE_PFX_BASE64,
            pfx_password=config.CERT_NE_PFX_PASSWORD, label="NE",
        )
        detalhes.update(ne)
    return detalhes


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def _get_nested(raw: dict, *chaves: str) -> Any:
    for chave in chaves:
        v: Any = raw
        for parte in chave.split("."):
            if not isinstance(v, dict):
                v = None
                break
            v = v.get(parte)
        if v is not None and v != "":
            return v
    return ""


def _extrair_campos(numero: str, raw: dict) -> dict:
    requerente = (
        raw.get("requerente") or raw.get("importador") or
        raw.get("exportador") or raw.get("solicitante") or {}
    )
    mercadoria = (
        raw.get("mercadoria") or raw.get("produto") or
        raw.get("item") or raw.get("dadosMercadoria") or {}
    )
    situacao_obj   = raw.get("situacao") if isinstance(raw.get("situacao"), dict) else {}
    pais_obj       = raw.get("paisDestino") or raw.get("pais") or raw.get("paisDeDestino") or {}
    porto_obj      = raw.get("portoEmbarque") or raw.get("porto") or raw.get("localEmbarque") or {}
    orgao_obj      = raw.get("orgaoAnuente") or raw.get("orgao") or {}
    transporte_obj = raw.get("modoTransporte") or raw.get("transporte") or {}

    cnpj = (
        _get_nested(requerente, "cpfCnpj", "cnpj", "cpf") or
        _get_nested(raw, "cpfCnpj", "cnpj")
    )
    nome = (
        _get_nested(requerente, "nome", "razaoSocial", "nomeEmpresa") or
        _get_nested(raw, "nomeRequerente", "nomeEmpresa")
    )

    quantidade_raw = (
        _get_nested(mercadoria, "quantidadeAutorizada", "quantidade", "qtd") or
        _get_nested(raw, "quantidadeAutorizada", "quantidade")
    )
    try:
        quantidade = float(quantidade_raw) if quantidade_raw else 0.0
    except (ValueError, TypeError):
        quantidade = 0.0

    qtd_utilizada_raw = (
        _get_nested(mercadoria, "quantidadeUtilizada", "qtdUtilizada") or
        _get_nested(raw, "quantidadeUtilizada", "qtdUtilizada")
    )
    try:
        qtd_utilizada = float(qtd_utilizada_raw) if qtd_utilizada_raw else 0.0
    except (ValueError, TypeError):
        qtd_utilizada = 0.0

    sit_id   = (
        _get_nested(situacao_obj, "id", "codigo") or
        (raw.get("situacao") if isinstance(raw.get("situacao"), str) else "") or ""
    )
    sit_desc = _get_nested(situacao_obj, "descricao", "nome") or ""

    pais_desc    = _get_nested(pais_obj, "descricao", "nome") or (pais_obj if isinstance(pais_obj, str) else "")
    porto_desc   = _get_nested(porto_obj, "descricao", "nome") or (porto_obj if isinstance(porto_obj, str) else "")
    orgao_desc   = _get_nested(orgao_obj, "descricao", "nome", "sigla") or (orgao_obj if isinstance(orgao_obj, str) else "")
    modal_desc   = _get_nested(transporte_obj, "descricao", "nome") or (transporte_obj if isinstance(transporte_obj, str) else "")

    modelo = _get_nested(raw, "codigoModelo", "modelo", "tipoLpco")

    # Saldo = autorizado - utilizado (se ambos disponíveis)
    saldo_raw = _get_nested(mercadoria, "saldo", "quantidadeSaldo") or _get_nested(raw, "saldo", "quantidadeSaldo")
    try:
        saldo = float(saldo_raw) if saldo_raw else (quantidade - qtd_utilizada if qtd_utilizada else None)
    except (ValueError, TypeError):
        saldo = None

    return {
        "numero_lpco":        numero,
        "cnpj":               cnpj,
        "nome_empresa":       nome,
        "uf":                 _get_nested(requerente, "uf", "estado", "siglaUf") or _get_nested(raw, "uf"),
        "municipio":          _get_nested(requerente, "municipio", "cidade") or _get_nested(raw, "municipio"),
        "codigo_modelo":      modelo,
        "tipo_lpco":          _MODELOS.get(str(modelo), modelo),
        "orgao_anuente":      orgao_desc,
        "ncm":                _get_nested(mercadoria, "ncm", "codigoNcm") or _get_nested(raw, "ncm", "codigoNcm"),
        "descricao_produto":  _get_nested(mercadoria, "descricao", "descricaoNcm", "nome") or _get_nested(raw, "descricaoProduto"),
        "quantidade":         quantidade,
        "qtd_utilizada":      qtd_utilizada,
        "saldo":              saldo,
        "unidade":            _get_nested(mercadoria, "unidade", "siglaUnidade") or _get_nested(raw, "unidade"),
        "pais_destino":       pais_desc,
        "porto_embarque":     porto_desc,
        "embarcacao":         _get_nested(raw, "embarcacao", "nomeEmbarcacao", "navio"),
        "modal_transporte":   modal_desc,
        "data_emissao":       _get_nested(raw, "dataEmissao", "dataAbertura", "dtEmissao"),
        "data_validade":      _get_nested(raw, "dataValidade", "validade", "dtValidade"),
        "situacao_id":        sit_id.upper() if sit_id else "",
        "situacao_desc":      sit_desc,
        "numero_processo":    _get_nested(raw, "numeroProcesso", "processo", "numProcesso"),
        "numero_di_due":      _get_nested(raw, "numeroDUE", "numeroDI", "numDUE", "numDI"),
    }


# ---------------------------------------------------------------------------
# Helpers de formatação Excel
# ---------------------------------------------------------------------------

def _aplicar_cabecalho(ws: Any, titulos: list[str], cor_hex: str = "1F4E79") -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(titulos)
    fill  = PatternFill("solid", fgColor=cor_hex)
    fonte = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = fonte
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _ajustar_colunas(ws: Any) -> None:
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        largura = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(largura + 4, 55)


def _linha_titulo_secao(ws: Any, texto: str, ncols: int) -> None:
    """Insere linha com fundo cinza como separador visual."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append([texto] + [""] * (ncols - 1))
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=min(ncols, ws.max_column or ncols))
    cell = ws.cell(row=row, column=1)
    cell.font      = Font(bold=True, size=12, color="1F4E79")
    cell.fill      = PatternFill("solid", fgColor="D9E1F2")
    cell.alignment = Alignment(horizontal="left")


def _top_n(contador: dict, n: int = 10) -> list[tuple]:
    return sorted(contador.items(), key=lambda x: -x[1])[:n]


# ---------------------------------------------------------------------------
# Aba 1 — Resumo por Cliente (LPCO)
# ---------------------------------------------------------------------------

def _aba_resumo_clientes(wb: Any, dados: list[dict]) -> None:
    from openpyxl.styles import Alignment
    ws = wb.active
    ws.title = "Resumo por Cliente (LPCO)"

    clientes: dict[str, dict] = defaultdict(lambda: {
        "nome": "", "uf": "", "municipio": "",
        "lpcos": set(), "total_kg": 0.0, "utilizado_kg": 0.0,
        "unidades": set(), "modelos": set(), "produtos": set(),
        "ncms": set(), "paises": set(), "portos": set(),
        "embarcacoes": set(), "modais": set(), "situacoes": defaultdict(int),
    })

    for d in dados:
        chave = d["cnpj"] or "N/I"
        c = clientes[chave]
        if d["nome_empresa"]:
            c["nome"] = d["nome_empresa"]
        if d["uf"]:
            c["uf"] = d["uf"]
        if d["municipio"]:
            c["municipio"] = d["municipio"]
        c["lpcos"].add(d["numero_lpco"])
        c["total_kg"] += d["quantidade"]
        if d["qtd_utilizada"]:
            c["utilizado_kg"] += d["qtd_utilizada"]
        for campo, dest in [
            ("unidade",          "unidades"),
            ("tipo_lpco",        "modelos"),
            ("descricao_produto","produtos"),
            ("ncm",              "ncms"),
            ("pais_destino",     "paises"),
            ("porto_embarque",   "portos"),
            ("embarcacao",       "embarcacoes"),
            ("modal_transporte", "modais"),
        ]:
            if d.get(campo):
                c[dest].add(d[campo])
        c["situacoes"][d["situacao_id"] or "SEM_INFO"] += 1

    _aplicar_cabecalho(ws, [
        "CNPJ / CPF", "Nome Empresa", "UF", "Município",
        "Total LPCOs", "Qtd. Autorizada (Total)", "Qtd. Utilizada (Total)", "Unidade",
        "Modelos (Tipo LPCO)", "NCMs", "Produtos / Espécies",
        "Países Destino", "Portos Embarque", "Embarcações", "Modais de Transporte",
        "Situações",
    ])

    for cnpj, c in sorted(clientes.items(), key=lambda x: -len(x[1]["lpcos"])):
        sits_str = "  ".join(f"{k}:{v}" for k, v in sorted(c["situacoes"].items()))
        util_pct = ""
        if c["total_kg"] and c["utilizado_kg"]:
            util_pct = f"{c['utilizado_kg']:.1f} ({100*c['utilizado_kg']/c['total_kg']:.0f}%)"
        ws.append([
            cnpj, c["nome"], c["uf"], c["municipio"],
            len(c["lpcos"]),
            round(c["total_kg"], 3) if c["total_kg"] else "",
            util_pct,
            " | ".join(sorted(c["unidades"])),
            " | ".join(sorted(c["modelos"])),
            " | ".join(sorted(c["ncms"])),
            " | ".join(sorted(c["produtos"])),
            " | ".join(sorted(c["paises"])),
            " | ".join(sorted(c["portos"])),
            " | ".join(sorted(c["embarcacoes"])),
            " | ".join(sorted(c["modais"])),
            sits_str,
        ])

    _ajustar_colunas(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Aba 2 — Detalhe LPCO
# ---------------------------------------------------------------------------

def _aba_detalhe_lpco(wb: Any, dados: list[dict]) -> None:
    from openpyxl.styles import PatternFill
    ws = wb.create_sheet("Detalhe LPCO")

    _aplicar_cabecalho(ws, [
        "Número LPCO", "CNPJ / CPF", "Nome Empresa", "UF", "Município",
        "Modelo", "Tipo", "Órgão Anuente",
        "NCM", "Produto / Espécie",
        "Qtd. Autorizada", "Qtd. Utilizada", "Saldo", "Unidade",
        "País Destino", "Porto Embarque", "Embarcação", "Modal Transporte",
        "Data Emissão", "Data Validade",
        "Número Processo", "Número DUE/DI",
        "Situação",
    ])

    for d in sorted(dados, key=lambda x: (x.get("cnpj", ""), x.get("numero_lpco", ""))):
        sit_label = d["situacao_id"]
        if d["situacao_desc"]:
            sit_label = f"{d['situacao_id']} — {d['situacao_desc']}"

        ws.append([
            d["numero_lpco"], d["cnpj"], d["nome_empresa"], d["uf"], d["municipio"],
            d["codigo_modelo"], d["tipo_lpco"], d["orgao_anuente"],
            d["ncm"], d["descricao_produto"],
            d["quantidade"] or "", d["qtd_utilizada"] or "", d["saldo"] if d["saldo"] is not None else "", d["unidade"],
            d["pais_destino"], d["porto_embarque"], d["embarcacao"], d["modal_transporte"],
            d["data_emissao"], d["data_validade"],
            d["numero_processo"], d["numero_di_due"],
            sit_label,
        ])

        cor = _CORES_SITUACAO.get(d["situacao_id"].upper() if d["situacao_id"] else "", "")
        if cor:
            fill = PatternFill("solid", fgColor=cor)
            for cell in ws[ws.max_row]:
                cell.fill = fill

    _ajustar_colunas(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Aba 3 — Resumo Mercado (DUEs)
# ---------------------------------------------------------------------------

def _aba_resumo_dues(wb: Any, dues: list[dict]) -> None:
    ws = wb.create_sheet("Resumo Mercado (DUEs)")
    _aplicar_cabecalho(ws, [
        "CNPJ Exportador", "Nome Exportador",
        "Total DUEs", "Total Eventos",
        "Peso Líquido Total (kg)", "Valor FOB Total (USD)",
        "NCMs / Produtos", "Países Destino",
        "Portos Embarque", "Embarcações", "Tipos de Evento",
    ], cor_hex="375623")

    mercado: dict[str, dict] = defaultdict(lambda: {
        "nome": "", "dues": set(), "eventos": 0,
        "peso_kg": 0.0, "fob_usd": 0.0,
        "ncms": set(), "paises": set(), "portos": set(),
        "embarcacoes": set(), "tipos": set(),
    })

    for d in dues:
        chave = d.get("exportador_cnpj") or d.get("numero_due", "N/I")[:12]
        m = mercado[chave]
        if d.get("exportador_nome"):
            m["nome"] = d["exportador_nome"]
        m["dues"].add(d["numero_due"])
        m["eventos"] += 1
        if d.get("peso_liquido_kg"):
            try:
                m["peso_kg"] += float(d["peso_liquido_kg"])
            except (TypeError, ValueError):
                pass
        if d.get("valor_fob_usd"):
            try:
                m["fob_usd"] += float(d["valor_fob_usd"])
            except (TypeError, ValueError):
                pass
        for campo, dest in [
            ("produto_ncm",    "ncms"),
            ("pais_destino",   "paises"),
            ("porto_embarque", "portos"),
            ("embarcacao",     "embarcacoes"),
            ("tipo_evento",    "tipos"),
        ]:
            if d.get(campo):
                m[dest].add(str(d[campo]))

    for cnpj, m in sorted(mercado.items(), key=lambda x: -len(x[1]["dues"])):
        ws.append([
            cnpj,
            m["nome"] or "(aguardando detalhe API)",
            len(m["dues"]),
            m["eventos"],
            round(m["peso_kg"], 1) if m["peso_kg"] else "",
            round(m["fob_usd"], 2) if m["fob_usd"] else "",
            " | ".join(sorted(m["ncms"])),
            " | ".join(sorted(m["paises"])),
            " | ".join(sorted(m["portos"])),
            " | ".join(sorted(m["embarcacoes"])),
            " | ".join(sorted(m["tipos"])),
        ])

    _ajustar_colunas(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Aba 4 — Detalhe DUEs
# ---------------------------------------------------------------------------

def _aba_detalhe_dues(wb: Any, dues: list[dict]) -> None:
    ws = wb.create_sheet("Detalhe DUEs")
    _aplicar_cabecalho(ws, [
        "Número DUE", "Data Evento", "Data Recebido",
        "Tipo Evento", "Descrição do Evento",
        "Exportador CNPJ", "Exportador Nome",
        "NCM", "Produto / Espécie",
        "Peso Líquido (kg)", "Peso Bruto (kg)", "Valor FOB (USD)",
        "País Destino", "Porto Embarque", "Embarcação",
        "RUC",
    ], cor_hex="375623")

    for d in sorted(dues, key=lambda x: x.get("data_evento", "")):
        ws.append([
            d.get("numero_due", ""),
            d.get("data_evento", ""),
            d.get("data_recebido", ""),
            d.get("tipo_evento", ""),
            d.get("descricao_evento", ""),
            d.get("exportador_cnpj", ""),
            d.get("exportador_nome", ""),
            d.get("produto_ncm", ""),
            d.get("produto_desc", ""),
            d.get("peso_liquido_kg") or "",
            d.get("peso_bruto_kg") or "",
            d.get("valor_fob_usd") or "",
            d.get("pais_destino", ""),
            d.get("porto_embarque", ""),
            d.get("embarcacao", ""),
            d.get("ruc", ""),
        ])

    _ajustar_colunas(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Aba 5 — Análise de Mercado (rankings combinados)
# ---------------------------------------------------------------------------

def _aba_analise_mercado(wb: Any, dados_lpco: list[dict], dues: list[dict]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet("Análise de Mercado")

    ncols = 4

    def _titulo(texto: str) -> None:
        _linha_titulo_secao(ws, texto, ncols)

    def _cab(*cols: str) -> None:
        _aplicar_cabecalho(ws, list(cols), cor_hex="4472C4")

    def _separador() -> None:
        ws.append([""] * ncols)

    # -----------------------------------------------------------------------
    # LPCO — Rankings
    # -----------------------------------------------------------------------
    _titulo("■ LPCO — Rankings do Período")
    _separador()

    # Top exportadores por número de LPCOs
    _cab("Rank", "Exportador", "CNPJ", "Qtd. LPCOs")
    cont_exp: dict[str, dict] = defaultdict(lambda: {"nome": "", "count": 0})
    for d in dados_lpco:
        k = d["cnpj"] or "N/I"
        cont_exp[k]["nome"] = d["nome_empresa"] or cont_exp[k]["nome"]
        cont_exp[k]["count"] += 1
    for rank, (cnpj, info) in enumerate(_top_n({k: v["count"] for k, v in cont_exp.items()}, 15), 1):
        ws.append([rank, cont_exp[cnpj]["nome"], cnpj, info])

    _separador()

    # Top países destino (LPCO)
    _cab("Rank", "País Destino", "", "Qtd. LPCOs")
    cont_pais: dict[str, int] = defaultdict(int)
    for d in dados_lpco:
        if d["pais_destino"]:
            cont_pais[d["pais_destino"]] += 1
    for rank, (pais, qtd) in enumerate(_top_n(cont_pais, 15), 1):
        ws.append([rank, pais, "", qtd])

    _separador()

    # Top produtos/NCMs (LPCO)
    _cab("Rank", "NCM", "Produto / Espécie", "Qtd. LPCOs")
    cont_ncm: dict[str, dict] = defaultdict(lambda: {"desc": "", "count": 0})
    for d in dados_lpco:
        k = d["ncm"] or "N/I"
        cont_ncm[k]["desc"] = d["descricao_produto"] or cont_ncm[k]["desc"]
        cont_ncm[k]["count"] += 1
    for rank, (ncm, info) in enumerate(_top_n({k: v["count"] for k, v in cont_ncm.items()}, 15), 1):
        ws.append([rank, ncm, cont_ncm[ncm]["desc"], info])

    _separador()

    # Top portos de embarque (LPCO)
    _cab("Rank", "Porto de Embarque", "", "Qtd. LPCOs")
    cont_porto: dict[str, int] = defaultdict(int)
    for d in dados_lpco:
        if d["porto_embarque"]:
            cont_porto[d["porto_embarque"]] += 1
    for rank, (porto, qtd) in enumerate(_top_n(cont_porto, 10), 1):
        ws.append([rank, porto, "", qtd])

    _separador()

    # Distribuição por situação (LPCO)
    _cab("Situação", "Descrição", "", "Qtd. LPCOs")
    cont_sit: dict[str, int] = defaultdict(int)
    for d in dados_lpco:
        cont_sit[d["situacao_id"] or "SEM_INFO"] += 1
    for sit, qtd in sorted(cont_sit.items(), key=lambda x: -x[1]):
        ws.append([sit, "", "", qtd])

    _separador()

    # -----------------------------------------------------------------------
    # DUE — Rankings de Mercado
    # -----------------------------------------------------------------------
    if dues:
        _titulo("■ DUEs — Pesquisa de Mercado (eventos recebidos via webhook)")
        _separador()

        # Top exportadores por DUEs
        _cab("Rank", "Exportador / CNPJ", "Nome", "Total DUEs")
        cont_due_exp: dict[str, dict] = defaultdict(lambda: {"nome": "", "dues": set()})
        for d in dues:
            k = d.get("exportador_cnpj") or d.get("numero_due", "N/I")[:12]
            cont_due_exp[k]["nome"] = d.get("exportador_nome", "") or cont_due_exp[k]["nome"]
            cont_due_exp[k]["dues"].add(d["numero_due"])
        for rank, (cnpj, info) in enumerate(
            sorted(cont_due_exp.items(), key=lambda x: -len(x[1]["dues"]))[:15], 1
        ):
            ws.append([rank, cnpj, info["nome"] or "(sem detalhe)", len(info["dues"])])

        _separador()

        # Top países destino (DUE)
        _cab("Rank", "País Destino (DUE)", "", "Qtd. DUEs")
        cont_due_pais: dict[str, int] = defaultdict(int)
        for d in dues:
            if d.get("pais_destino"):
                cont_due_pais[d["pais_destino"]] += 1
        for rank, (pais, qtd) in enumerate(_top_n(cont_due_pais, 15), 1):
            ws.append([rank, pais, "", qtd])

        _separador()

        # Top tipos de evento (DUE)
        _cab("Tipo Evento DUE", "Descrição", "", "Qtd. Eventos")
        cont_tipo: dict[str, int] = defaultdict(int)
        for d in dues:
            cont_tipo[d.get("tipo_evento") or "N/I"] += 1
        for tipo, qtd in sorted(cont_tipo.items(), key=lambda x: -x[1]):
            ws.append([tipo, "", "", qtd])

        _separador()

    # -----------------------------------------------------------------------
    # Resumo geral
    # -----------------------------------------------------------------------
    _titulo("■ Totais Gerais")
    _separador()
    ws.append(["Total LPCOs no relatório:", len(dados_lpco), "", ""])
    ws.append(["Total exportadores únicos (LPCO):", len(cont_exp), "", ""])
    ws.append(["Total DUEs recebidas (período):", len(set(d["numero_due"] for d in dues)) if dues else 0, "", ""])
    ws.append(["Total exportadores únicos (DUE):", len(cont_due_exp) if dues else 0, "", ""])

    _ajustar_colunas(ws)


# ---------------------------------------------------------------------------
# Aba 6 — Legenda
# ---------------------------------------------------------------------------

def _aba_legenda(wb: Any) -> None:
    from openpyxl.styles import Font, PatternFill
    ws = wb.create_sheet("Legenda")
    ws.append(["Cor", "Situação LPCO", "Descrição"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    legenda = [
        ("C6EFCE", "DEFERIDO",      "LPCO aprovado pelo órgão anuente"),
        ("FFC7CE", "INDEFERIDO",     "LPCO negado pelo órgão anuente"),
        ("FFEB9C", "EM_ANALISE",     "Em análise pelo órgão anuente"),
        ("DDEBF7", "EM_VERIFICACAO", "Aguardando verificação/documentação"),
        ("D9D9D9", "CANCELADO",      "LPCO cancelado"),
        ("F2DCDB", "VENCIDO",        "LPCO com prazo de validade expirado"),
        ("E2EFDA", "SUSPENSAO",      "LPCO em suspensão"),
        ("FFFFFF", "Demais",         "Outras situações sem cor específica"),
    ]
    for cor, sit, desc in legenda:
        ws.append(["", sit, desc])
        ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=cor)
    from openpyxl.utils import get_column_letter
    for i, larg in enumerate([8, 25, 50], start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg


# ---------------------------------------------------------------------------
# Montagem final do workbook
# ---------------------------------------------------------------------------

def _gerar_excel_completo(dados_lpco: list[dict], dues: list[dict], periodo_label: str) -> bytes:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl não instalado — execute: pip install openpyxl")

    wb = openpyxl.Workbook()

    _aba_resumo_clientes(wb, dados_lpco)
    _aba_detalhe_lpco(wb, dados_lpco)
    if dues:
        _aba_resumo_dues(wb, dues)
        _aba_detalhe_dues(wb, dues)
    _aba_analise_mercado(wb, dados_lpco, dues)
    _aba_legenda(wb)

    wb.properties.description = f"Relatório Mercado Hevile — {periodo_label}"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Ponto de entrada do job semanal
# ---------------------------------------------------------------------------

def gerar_e_enviar_relatorio_semanal() -> None:
    """
    Varre TODOS os LPCOs acessíveis via SE e NE (sem filtro do banco),
    combina com DUEs acumuladas no período e envia Excel por email.
    Chamado pelo APScheduler toda sexta-feira às 14:00.
    """
    from email_service import enviar_relatorio_excel

    agora   = datetime.now()
    inicio  = agora - timedelta(days=7)
    periodo = f"{inicio.strftime('%d/%m/%Y')} a {agora.strftime('%d/%m/%Y')}"
    nome    = f"relatorio_mercado_{agora.strftime('%Y-%m-%d')}.xlsx"

    logger.info("=== Relatório semanal iniciando — período: %s ===", periodo)

    # 1. LPCOs via API
    detalhes   = _buscar_todos_detalhes()
    dados_lpco = [_extrair_campos(num, raw) for num, raw in detalhes.items()] if detalhes else []
    logger.info("LPCOs: %d total, %d com detalhe.", len(detalhes), sum(1 for v in detalhes.values() if v))

    # 2. DUEs acumuladas no banco (período da semana)
    dues = listar_dues(
        data_inicio=inicio.strftime("%Y-%m-%d"),
        data_fim=agora.strftime("%Y-%m-%d"),
    )
    logger.info("DUEs no período: %d", len(dues))

    if not dados_lpco and not dues:
        logger.error("Sem dados para o relatório — cancelado.")
        return

    # 3. Gera Excel
    try:
        excel_bytes = _gerar_excel_completo(dados_lpco, dues, periodo)
    except Exception as exc:
        logger.error("Erro ao gerar Excel: %s", exc)
        return

    # 4. Envia por email
    try:
        enviar_relatorio_excel(excel_bytes, nome, periodo)
        logger.info("=== Relatório enviado: %s ===", nome)
    except Exception as exc:
        logger.error("Erro ao enviar relatório: %s", exc)
